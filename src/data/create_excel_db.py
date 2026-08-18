import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json
import csv
import os

def create_full_excel_db():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_dir, "inteplast_products_database.xlsx")
    csv_path = os.path.join(base_dir, "products.csv")
    json_path = os.path.join(base_dir, "products.json")

    wb = openpyxl.Workbook()

    # Styling Token Definitions
    header_fill = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
    header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    
    sub_header_fill = PatternFill(start_color="00529B", end_color="00529B", fill_type="solid")
    sub_header_font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
    
    data_font = Font(name="微軟正黑體", size=10)
    bold_data_font = Font(name="微軟正黑體", size=10, bold=True)
    
    zebra_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # -------------------------------------------------------------
    # Sheet 1: 產品分類總覽 (Categories Summary)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "產品分類總覽"
    ws1.views.sheetView[0].showGridLines = True

    headers1 = [
        "產品編號", "分類識別碼 (ID)", "產品中文名稱", "產品英文名稱", 
        "產品圖片檔名/路徑", "特色標籤 (多個請用逗號分隔)", "核心亮點與防護優勢", 
        "產品詳細介紹文案", "專頁檔案連結"
    ]

    ws1.append(headers1)
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows1 = [
        [
            "001", "cat-can-liners", "清潔袋系列 (Can Liners)", "Can Liners & Trash Bags",
            "src/formosa_canliner_official.png", "連捲點斷, 單張抽取, 環保標章認證",
            "高強度防漏 · 商業機構與家用大容量",
            "包含連捲清潔袋、單張平裝抽取清潔袋與通過環保標章認證之 50% 再生塑膠環保清潔袋。選用台塑 PE 純料韌性極佳，底封堅固不易破裂滴漏。",
            "can-liners.html"
        ],
        [
            "002", "cat-draw-tape", "拉繩袋系列 (Draw-Tape Bags)", "Draw-Tape & Medical Bags",
            "src/Can-Liners-Draw-Tape-Draw-Tape-2.jpg", "一拉即束, 不髒手打包, 醫療隔離警示",
            "專利加長拉繩 · 輕鬆拎舉不黏手打包",
            "加長拉繩設計，雙手一拉自動提束包緊，打包過程衛生不髒手；並提供醫療等級高辨識度黃色與紅色傳染廢棄物隔離袋。",
            "draw-tape.html"
        ],
        [
            "003", "cat-heat-bags", "耐熱袋系列 (Heat-Resistant Bags)", "Foodservice Heat-Resistant Bags",
            "src/formosa_heat_bag_official.png", "食品級 PE, 平裝 / 捲裝, 高溫耐受",
            "嚴選台塑食品級原料 · 絕無塑化劑添加",
            "平裝與捲裝耐熱袋，適用於生鮮備料、小吃餐飲熱食分裝與高溫熱湯打包。通過國家食安標準審查，品質安全無虞。",
            "foodservice.html"
        ],
        [
            "004", "cat-zipper-bags", "密封包裝類 (Sealed & Zipper Bags)", "Sealed & Zipper Storage Bags",
            "src/ai_zipper_bag.png", "雙軌密實, 立體站立, 冷凍保鮮",
            "嚴密隔絕空氣濕氣 · 食材長效保鮮",
            "提供雙軌密封密實袋、立體可站立密實袋、冷凍保鮮袋與通用夾鏈袋，滿足家用廚房食材分裝與商業食品儲存鎖鮮需求。",
            "foodservice.html"
        ],
        [
            "005", "cat-specialty", "其他類 (Specialty Protection & Tapes)", "Specialty Protection & Tapes",
            "src/ai_specialty.png", "手套, 台塑遮蔽防塵膠帶",
            "專業個人防護與建築施工/裝潢養生防塵",
            "包含輕薄貼手防護手套，以及 550mm~3200mm 多種規格之台塑專利遮蔽防塵膠帶，適合建築施工、室內裝潢養生與日常防塵防護。",
            "specialty.html"
        ],
        [
            "006", "cat-scale-sheet", "Scale Sheet", "Scale Sheet",
            "src/scale sheet.png", "", "", "", "stretch-films.html"
        ]
    ]

    for row_idx, row_data in enumerate(rows1, 2):
        ws1.append(row_data)
        fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, len(row_data) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col_idx in [1, 2, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # -------------------------------------------------------------
    # Sheet 2: 完整容量與尺寸規格對照總表 (Complete Specs Matrix)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="容量尺寸規格對照總表")
    ws2.views.sheetView[0].showGridLines = True

    headers2 = [
        "產品編號", "所屬產品系列", "規格/容量/號數", "尺寸 (寬 × 長 cm / mm)", "數量 / 張數 / 卷長", "顏色 / 色系", "適用情境 / 特色說明"
    ]

    ws2.append(headers2)
    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = sub_header_fill
        cell.font = sub_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows2 = [
        # 001 清潔袋
        ["001", "清潔袋系列", "10L 特小", "40 × 50 cm", "100 張/捲", "透明", "家用廚房/小型垃圾桶"],
        ["001", "清潔袋系列", "15L 小", "43 × 56 cm", "72 張/捲", "透明 / 粉紅", "家用/辦公室通用"],
        ["001", "清潔袋系列", "20L 中", "53 × 63 cm", "54 張/捲", "透明 / 粉紅", "辦公室/通用型"],
        ["001", "清潔袋系列", "45L 大", "65 × 75 cm", "30 張/捲", "透明 / 粉紅", "熱銷推薦/家庭大容量"],
        ["001", "清潔袋系列", "70L 特大", "80 × 90 cm", "22 張/捲", "透明 / 黑色", "商業機構/餐廳"],
        ["001", "清潔袋系列", "90L 超大", "86 × 100 cm", "16 張/捲", "透明 / 黑色", "工業加厚/飯店"],
        ["001", "清潔袋系列", "125L 超大", "90 × 110 cm", "28 張/捲", "透明 / 黑色", "超大容積/戶外環境"],
        ["001", "清潔袋系列", "126L 超特大", "91 × 110 cm", "10 張/捲", "透明 / 黑色", "巨無霸/環境清潔"],

        # 002 拉繩袋 (依據台塑拉繩清潔袋與醫療袋官方包裝圖檔)
        ["002", "拉繩袋系列", "45L 大", "65 × 75 cm", "24張 / 經濟包", "黑 / 透明", "台塑拉繩清潔袋 / 雙向抽拉打包"],
        ["002", "拉繩袋系列", "70L 特大", "78 × 88 cm", "18張 / 經濟包", "黑 / 透明", "台塑拉繩清潔袋 / 雙向抽拉打包"],
        ["002", "拉繩袋系列", "90L 超大", "84 × 95 cm", "14張 / 經濟包", "黑 / 透明", "台塑拉繩清潔袋 / 雙向抽拉打包"],
        ["002", "拉繩袋系列", "90L 超大 (超量包)", "84 × 95 cm", "25張 / 超量包", "黑 / 透明", "台塑拉繩清潔袋 / 雙向抽拉打包"],
        ["002", "拉繩袋系列", "125L 超特大", "93 × 100 cm", "15張 / 經濟包", "黑 / 透明", "台塑拉繩清潔袋 / 雙向抽拉打包"],
        ["002", "拉繩袋系列", "130L 巨無霸", "94 × 102 cm", "20張 / 超量包", "黑 / 透明", "台塑拉繩清潔袋 / 雙向抽拉打包"],
        ["002", "拉繩袋系列", "90L 超大 (彩色版)", "84 × 95 cm", "25張 / 超量包", "藍 / 粉", "台塑拉繩清潔袋 / 馬卡龍雙色"],
        ["002", "拉繩袋系列", "90L 超大 (盒裝)", "84 × 95 cm", "80張 / 盒裝", "黑", "台塑拉繩清潔袋 / 抽取盒裝"],
        ["002", "拉繩袋系列", "125L 超特大 (盒裝)", "93 × 100 cm", "66張 / 盒裝", "黑", "台塑拉繩清潔袋 / 抽取盒裝"],

        # 002 醫療拉繩袋
        ["002", "拉繩袋系列", "醫療袋 8L 小", "39 × 40 cm", "40張 / 捲", "紅", "台塑拉繩醫療感染袋 / 感控隔離專用"],
        ["002", "拉繩袋系列", "醫療袋 20L 中", "52 × 55 cm", "22張 / 捲", "紅", "台塑拉繩醫療感染袋 / 感控隔離專用"],
        ["002", "拉繩袋系列", "醫療袋 50L 大", "69 × 78 cm", "12張 / 捲", "紅", "台塑拉繩醫療感染袋 / 感控隔離專用"],
        ["002", "拉繩袋系列", "醫療袋 70L 特大", "77 × 92 cm", "8張 / 捲", "紅", "台塑拉繩醫療感染袋 / 感控隔離專用"],
        ["002", "拉繩袋系列", "醫療袋 90L 超大", "95 × 84 cm", "12張 / 捲", "紅", "台塑拉繩醫療感染袋 / 感控隔離專用"],

        # 003 耐熱袋
        ["003", "耐熱袋系列", "四兩", "12.7 × 17.8 cm", "平裝 / 捲裝", "透明食品級", "醬料/小份量食材打包"],
        ["003", "耐熱袋系列", "半斤", "15.2 × 22.8 cm", "平裝 / 捲裝", "透明食品級", "小吃/醬料/生鮮分裝"],
        ["003", "耐熱袋系列", "1斤", "20.3 × 27.9 cm", "平裝 / 捲裝", "透明食品級", "餐飲熱食打包"],
        ["003", "耐熱袋系列", "2斤", "25.4 × 35.5 cm", "平裝 / 捲裝", "透明食品級", "外帶熱湯/生鮮肉品"],
        ["003", "耐熱袋系列", "3斤", "30.4 × 43.1 cm", "平裝 / 捲裝", "透明食品級", "大份量熱食打包"],
        ["003", "耐熱袋系列", "5斤", "35.5 × 50.8 cm", "平裝 / 捲裝", "透明食品級", "商業備料/高溫高壓打包"],

        # 004 密封包裝類
        ["004", "密封包裝類", "1號夾鏈袋", "5.0 × 7.0 cm", "每包100入", "透明", "小型飾品/藥品分裝"],
        ["004", "密封包裝類", "2號夾鏈袋", "6.0 × 8.5 cm", "每包100入", "透明", "零件/隨身藥包"],
        ["004", "密封包裝類", "3號夾鏈袋", "7.0 × 10.0 cm", "每包100入", "透明", "卡片/小物收納"],
        ["004", "密封包裝類", "4號夾鏈袋", "8.5 × 12.0 cm", "每包100入", "透明", "試用品/小文具"],
        ["004", "密封包裝類", "5號夾鏈袋", "10.0 × 14.0 cm", "每包100入", "透明", "食材分裝/小點心"],
        ["004", "密封包裝類", "6號夾鏈袋", "12.0 × 17.0 cm", "每包100入", "透明", "中型食材/口罩收納"],
        ["004", "密封包裝類", "7號夾鏈袋", "14.0 × 20.0 cm", "每包100入", "透明", "蔬果/冷凍食品"],
        ["004", "密封包裝類", "8號夾鏈袋", "17.0 × 24.0 cm", "每包100入", "透明", "A5文件/冷凍肉品"],
        ["004", "密封包裝類", "9號夾鏈袋", "20.0 × 28.0 cm", "每包100入", "透明", "服飾/大份量食材"],
        ["004", "密封包裝類", "10號夾鏈袋", "24.0 × 34.0 cm", "每包100入", "透明", "A4文件/大物件"],
        ["004", "密封包裝類", "11號夾鏈袋", "28.0 × 40.0 cm", "每包100入", "透明", "大件衣物/收納"],
        ["004", "密封包裝類", "12號夾鏈袋", "34.0 × 45.0 cm", "每包100入", "透明", "超大物件收納"],

        # 005 其他類
        ["005", "其他類", "手套 S", "手掌寬 ~8 cm", "每盒100入", "半透明", "個人衛生/餐飲備料"],
        ["005", "其他類", "手套 M", "手掌寬 ~9 cm", "每盒100入", "半透明", "個人衛生/餐飲備料"],
        ["005", "其他類", "手套 L", "手掌寬 ~10 cm", "每盒100入", "半透明", "個人衛生/餐飲備料"],
        ["005", "其他類", "手套 XL", "手掌寬 ~11 cm", "每盒100入", "半透明", "個人衛生/餐飲備料"],
        ["005", "其他類", "防塵膠帶 550mm", "550 mm × 25 y (22.8m)", "卷裝", "綠色膠帶+高密度膜", "建築施工/小面積家具遮蔽"],
        ["005", "其他類", "防塵膠帶 1100mm", "1100 mm × 25 y (22.8m)", "卷裝", "綠色膠帶+高密度膜", "裝潢養生/牆面家具遮蔽"],
        ["005", "其他類", "防塵膠帶 1500mm", "1500 mm × 25 y (22.8m)", "卷裝", "綠色膠帶+高密度膜", "裝潢養生/牆面門窗遮蔽"],
        ["005", "其他類", "防塵膠帶 2100mm", "2100 mm × 25 y (22.8m)", "卷裝", "綠色膠帶+高密度膜", "門窗/大型家具大面積遮蔽"],
        ["005", "其他類", "防塵膠帶 2700mm", "2700 mm × 25 y (22.8m)", "卷裝", "綠色膠帶+高密度膜", "工地工程/超大面積遮蔽"],
        ["005", "其他類", "防塵膠帶 3200mm", "3200 mm × 25 y (22.8m)", "卷裝", "綠色膠帶+高密度膜", "工程施工/最大面積養生遮蔽"],

        # 006 Scale Sheet
        ["006", "Scale Sheet", "Scale Sheet", "專案洽詢", "專案對接", "專案客製", "如有需求請聯繫專員專人對接"]
    ]

    for row_idx, row_data in enumerate(rows2, 2):
        ws2.append(row_data)
        fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, len(row_data) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col_idx in [1, 3, 4, 5, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust Column Widths & Heights
    for ws in [ws1, ws2]:
        ws.row_dimensions[1].height = 28
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                cell_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                if cell_len > max_len:
                    max_len = cell_len
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save(excel_path)
    print(f"Successfully generated full specs Excel workbook: {excel_path}")

    # Build products.csv
    csv_rows = []
    # Build quick_specs dictionary per category code
    cat_specs_map = {}
    for r in rows2:
        code = r[0]
        size = r[2]
        dim = r[3]
        if code not in cat_specs_map:
            cat_specs_map[code] = []
        if size and dim and size != "Scale Sheet":
            cat_specs_map[code].append(f"{size}: {dim}")

    for r in rows1:
        code = r[0]
        specs_str = "; ".join(cat_specs_map.get(code, []))
        csv_rows.append({
            "id": r[1],
            "code": r[0],
            "title_tw": r[2],
            "title_en": r[3],
            "image": r[4],
            "badges": r[5],
            "highlights_tw": r[6],
            "desc_tw": r[7],
            "page_url": r[8],
            "quick_specs": specs_str
        })

    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        fieldnames = ["id", "code", "title_tw", "title_en", "image", "badges", "highlights_tw", "desc_tw", "page_url", "quick_specs"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(csv_rows)

    print(f"Successfully updated CSV database: {csv_path}")

if __name__ == "__main__":
    create_full_excel_db()
